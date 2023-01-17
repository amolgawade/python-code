import openpyxl
from openpyxl.styles import PatternFill


def getrowsCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def getcolumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column

def readData(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,columnno).value = data
    workbook.save(file)

def fillredcolor(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redfill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    sheet.cell(rownum,columnno).fill = redfill
    workbook.save(file)

def fillgreencolor(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    sheet.cell(rownum,columnno).fill = greenfill
    workbook.save(file)



