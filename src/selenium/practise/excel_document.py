import openpyxl

file = r"E:\code\python_practise\src\selenium\textcase1.xlsx"
work = openpyxl.load_workbook(file)
sheet = work['scenarios ']

rows = sheet.max_row
column = sheet.max_column

for r in range(1, rows + 1):
    for c in range(1, column + 1):
        print(sheet.cell(r, c).value, end="         ")
    print()
