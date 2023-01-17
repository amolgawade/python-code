import openpyxl

file = r"E:\code\python_practise\src\selenium\textcase1.xlsx"
work = openpyxl.load_workbook(file)
sheet = work.active

for r in range(1, 6):
    for c in range(1,4):
        sheet.cell(r,c).value='welcome'

sheet.cell(6, 1).value = "sr.no"
sheet.cell(6, 2).value = "name"
sheet.cell(6, 3).value = "designation"

sheet.cell(7, 1).value = "1"
sheet.cell(7, 2).value = "amol"
sheet.cell(7, 3).value = " BE"

sheet.cell(8, 1).value = "2"
sheet.cell(8, 2).value = "KANCHAN"
sheet.cell(8, 3).value = "BE"

sheet.cell(9, 1).value = "3"
sheet.cell(9, 2).value = "URVI"
sheet.cell(9, 3).value = "NURSSARY"

work.save(file)